from typing import Mapping
from openpyxl import Workbook
from openpyxl import load_workbook


def verify_file_suffix(check_str, target_str=['xls','']):
    assert type(check_str) is str
    assert type(target_str) is str and target_str != ''
    str(check_str).rfind('.')[0] == target_str

def write_new_workbook(filename, data):
    assert filename is not None
    assert type(filename) == str

    wb = Workbook()
    ws = wb.active

    # write data into it
    ws['A1'] = 42
    ws.append([1, 2, 3])

    # save excel file
    wb.save(filename)


def read_workbook(filename): 
    assert type(filename) is str
    assert filename != ''
    load_workbook('filename')
    


# test func
